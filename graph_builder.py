import os
import openpyxl
import networkx as nx
import matplotlib.pyplot as plt
import io
from openpyxl.drawing.image import Image
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

import re


def normalize_node_label(text):
    """
    Sanitizes node names.
    - Strips []():.
    - Replaces _ with space
    - Title Case
    - Filters N/A, Unknown, short
    """
    if not text:
        return None
        
    # Replace separators with space
    text = text.replace("_", " ").replace("-", " ")
    
    # Remove specific punctuation chars
    # Stripping []():. 
    import re
    text = re.sub(r'[\[\]\(\):.]', '', text)
    
    # Extra whitespace cleanup
    text = " ".join(text.split())
    
    # Title Case
    text = text.title()
    
    # Filtering
    if len(text) < 2:
        return None
    if text.lower() in ["n/a", "unknown", "none"]:
        return None
        
    return text

def extract_terms_from_sheet(sheet):
    """
    Extracts top technical/concept terms from a paper sheet.
    Priority: Glossary Terms -> Central Variables -> Title Keywords
    Returns a set of normalized (Title Case) terms.
    """
    terms = set()
    
    glossary_text = ""
    variables_text = ""
    
    # Scan rows for keys
    for row in sheet.iter_rows(min_row=1, max_row=30, min_col=1, max_col=2):
        key_cell = row[0]
        val_cell = row[1]
        if not key_cell.value:
            continue
            
        key = str(key_cell.value).strip()
        val = str(val_cell.value) if val_cell.value else ""
        
        if key == "Glossary":
            glossary_text = val
        elif key in ["Central Independent Variables", "Central Dependent Variables"]:
            variables_text += " " + val

    # Logic 1: Glossary Parsing
    import ast
    
    # Helper to clean and add
    def add_term(t):
        norm = normalize_node_label(t)
        if norm:
            terms.add(norm)

    # Try parsing as JSON/List repr first (if saved as raw list)
    # OR if saved as "Term: Definition", parse lines.
    
    # Case A: "Term: Definition\nTerm2: Def2" (Formatted String)
    if ":" in glossary_text and "\n" in glossary_text:
        lines = glossary_text.split('\n')
        for line in lines:
            if ":" in line:
                # Remove bullets if present
                clean_line = line.replace("•", "").strip()
                term = clean_line.split(":")[0]
                add_term(term)

    # Case B: Raw List/JSON "[{...}]" (If not formatted yet)
    elif glossary_text.strip().startswith("["):
        try:
            glossary_list = ast.literal_eval(glossary_text)
            if isinstance(glossary_list, list):
                for item in glossary_list:
                    if isinstance(item, dict):
                        term = item.get("Term", item.get("term", ""))
                        add_term(term)
        except:
            pass

    # Logic 2: Variables Parsing (Backup)
    if len(terms) < 3 and variables_text:
        parts = [p.strip() for p in variables_text.replace(" and ", ",").split(',')]
        for p in parts:
             add_term(p)

    return list(terms)[:8] # Increased limit slightly

def get_keywords(text):
    """Splits text into meaningful keywords (excluding stop words)."""
    stop_words = {"the", "of", "and", "in", "for", "a", "an", "to", "with", "on", "at", "by", "from"}
    words = text.lower().split()
    return {w for w in words if w not in stop_words and len(w) > 2}

def update_workbook_with_graph(wb):
    """
    Updates the provided workbook object with a Knowledge Graph sheet.
    Does NOT save the file (caller must save).
    """
    G = nx.Graph()
    
    # Iterate sheets
    skip_sheets = ["Dashboard", "Knowledge Graph"]
    
    paper_nodes = []
    concept_nodes = []
    
    logging.info("Building Knowledge Graph from workbook data...")
    
    # 1. Collect all Nodes
    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
            
        sheet = wb[sheet_name]
        
        # Add Paper Node
        G.add_node(sheet_name, type="paper")
        paper_nodes.append(sheet_name)
        
        # Extract Terms
        terms = extract_terms_from_sheet(sheet)
        
        # Add Concept Nodes & Connect to Paper (Direct Link)
        for term in terms:
            if not G.has_node(term):
                G.add_node(term, type="concept", keywords=get_keywords(term))
                concept_nodes.append(term)
            
            # Paper -> Concept (Strongest link, weight 5)
            G.add_edge(sheet_name, term, weight=5)

    if len(paper_nodes) == 0:
        logging.warning("No papers found to graph.")
        return

    # 2. Fuzzy Connectivity (Concept <-> Concept)
    # Link concepts that share keywords
    import itertools
    
    # Limit combinations if too many nodes to prevent N^2 explosion?
    # For < 500 nodes, N^2 is fine.
    
    for term_a, term_b in itertools.combinations(concept_nodes, 2):
        kw_a = G.nodes[term_a].get("keywords", set())
        kw_b = G.nodes[term_b].get("keywords", set())
        
        # Intersection
        common = kw_a & kw_b
        weight = len(common)
        
        if weight > 0:
            # Create edge with weight
            G.add_edge(term_a, term_b, weight=weight)


    # --- Visualization ---
    
    # Obsidian Theme Colors
    BG_COLOR = "#202020"
    PAPER_COLOR = "#FF4500" 
    CONCEPT_COLOR = "#00BFFF" 
    EDGE_COLOR = "#505050"
    TEXT_COLOR = "black" 
    
    # Increase DPI and figure size for better quality
    # Use fixed large size as requested
    # Increase DPI and figure size for better quality
    # Use fixed huge size as requested (40x40)
    plt.figure(figsize=(40, 40), facecolor=BG_COLOR, dpi=150)
    ax = plt.gca()
    ax.set_facecolor(BG_COLOR)
    
    # Dynamic Layout Physics: Spring Layout with High Repulsion
    import textwrap
    import math
    from adjustText import adjust_text
    
    # k = Optimal distance. High k = more repulsion.
    # 15.0 / sqrt(N) pushes nodes MASSIVELY apart (Supercharge Repulsion)
    num_nodes = len(G.nodes())
    k_val = 15.0 / math.sqrt(num_nodes) if num_nodes > 0 else 2.0
    
    # Scale=10 expands coordinate system significantly to handle the high repulsion
    # Iterations=300 allows the physics engine long enough to push everything apart
    pos = nx.spring_layout(G, k=k_val, iterations=300, seed=42, scale=10)
    
    # Sizes & Styles
    node_sizes = []
    node_colors = []
    
    # Calculate sizes based on Degree Centrality
    for node in G.nodes():
        degree = G.degree(node)
        
        # Papers are Hubs? No, usually Concepts connect many papers.
        # Regardless, use degree for size.
        
        # Base Size 3000, + 1000 per partial degree
        # Let's use the formula: Size = 3000 + (Degree * 1000)
        size = 3000 + (degree * 1000)
        node_sizes.append(size)

        if G.nodes[node].get("type") == "paper":
            node_colors.append(PAPER_COLOR)
        else:
            node_colors.append(CONCEPT_COLOR)
            
    # Draw Nodes with Borders
    nodes = nx.draw_networkx_nodes(G, pos, 
                          node_size=node_sizes, 
                          node_color=node_colors, 
                          alpha=1.0, 
                          linewidths=3, # Thicker border for big nodes
                          edgecolors="white",
                          ax=ax)
    
    # Draw Edges 
    edges = G.edges()
    weights = [G[u][v].get('weight', 1) for u, v in edges]
    widths = [w * 1.0 for w in weights] # Slightly thicker edges
    
    nx.draw_networkx_edges(G, pos, 
                           edgelist=edges, 
                           width=widths, 
                           edge_color=EDGE_COLOR, 
                           alpha=0.5, 
                           arrows=True, # Required for connectionstyle
                           connectionstyle="arc3,rad=0.1", 
                           ax=ax)
    
    # Draw Labels with Dynamic Font Size
    texts = []
    for node, (x, y) in pos.items():
        degree = G.degree(node)
        
        # Font Size Calculation: 8 + (Degree * 1.2)
        font_size = 8 + (degree * 1.2)
        # Cap max font size purely for sanity (e.g. 24)
        if font_size > 24: 
            font_size = 24
        
        # Wrap text - loosen wrapping for larger fonts? Keep 15 for now.
        label_text = "\n".join(textwrap.wrap(str(node), width=15))
        
        # Create text object with dynamic size
        t = ax.text(x, y, label_text, 
                    fontsize=font_size, 
                    fontweight="bold", 
                    color=TEXT_COLOR,
                    fontfamily="sans-serif",
                    ha='center', va='center',
                    bbox=dict(boxstyle="round,pad=0.3", facecolor="white", edgecolor="none", alpha=0.6))
        texts.append(t)
        
    # Run the physics solver for labels
    if texts:
        logging.info("Adjusting label positions (this may take a moment)...")
        # Increase arrows to make label movement clearer
        # Add 'add_objects=[nodes]' so text bounces off the balls
        adjust_text(texts, 
                    add_objects=[nodes], 
                    arrowprops=dict(arrowstyle='-', color='gray', lw=1.0),
                    expand_points=(2.0, 2.0))
    
    plt.axis('off')
    plt.tight_layout()
    
    # Save to Buffer
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', facecolor=BG_COLOR, edgecolor='none', dpi=200)
    img_buffer.seek(0)
    plt.close()

    
    # --- Excel Integration ---
    GRAPH_SHEET_NAME = "Knowledge Graph"
    
    if GRAPH_SHEET_NAME in wb.sheetnames:
        ws = wb[GRAPH_SHEET_NAME]
        ws.delete_rows(1, 100) # Clear
        # Clear images? Hard to clear images easily in openpyxl without deleting sheet
        del wb[GRAPH_SHEET_NAME]
        ws = wb.create_sheet(GRAPH_SHEET_NAME, 0)
    else:
        ws = wb.create_sheet(GRAPH_SHEET_NAME, 0)
        
    # Move to index 0 (or 1 if dashboard is 0? User said "at Index 0 and desplace the rests to right")
    # If Dashboard is at 0, this will take 0 and push Dashboard to 1?
    # User said "Dashboard" as first tab in previous prompt.
    # Now says "Knowledge Graph" at Index 0.
    # I'll put Graph at 0, Dashboard at 1. Or maybe Graph at 1?
    # "at Index 0". I will obey.
    if wb.sheetnames[0] != GRAPH_SHEET_NAME:
        wb.move_sheet(ws, offset=-wb.index(ws))

    # Insert Image
    img = Image(img_buffer)
    img.anchor = 'A1'
    ws.add_image(img)
    
    logging.info(f"Knowledge Graph inserted into '{GRAPH_SHEET_NAME}' sheet.")

if __name__ == "__main__":
    # Test
    # build_graph_from_excel("Paper_Analysis_Results.xlsx")
    pass
