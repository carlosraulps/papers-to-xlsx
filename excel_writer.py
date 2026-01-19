import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import re

def clean_sheet_name(raw_name):
    """
    Cleans the sheet name to ensure it's valid for Excel.
    Excel sheet names: max 31 chars, no [] : * ? / \
    """
    # Remove invalid characters
    invalid_chars = r'[\[\]:*?/\\]'
    name = re.sub(invalid_chars, '', raw_name)
    # Truncate to 31 chars
    return name[:31]

def get_unique_sheet_name(workbook, base_name):
    """
    Generates a unique sheet name by appending a counter if specific name exists.
    """
    name = clean_sheet_name(base_name)
    if name not in workbook.sheetnames:
        return name
    
    count = 2
    while True:
        new_name = f"{name}_{count}"
        if len(new_name) > 31:
            # If counting makes it too long, truncate base further
            truncated_len = 31 - len(str(count)) - 1
            new_name = f"{name[:truncated_len]}_{count}"
            
        if new_name not in workbook.sheetnames:
            return new_name
        count += 1

def load_or_create_workbook(filename):
    """Loads an existing workbook or creates a new one if it doesn't exist or is corrupted."""
    import zipfile
    import os
    import shutil
    
    try:
        wb = openpyxl.load_workbook(filename)
        return wb
    except (FileNotFoundError):
        # Normal case for new run
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        return wb
    except (KeyError, zipfile.BadZipFile):
        # File is corrupted (e.g. '[Content_Types].xml' missing)
        print(f"WARNING: The Excel file {filename} is corrupted. Backing up and creating a new one.")
        if os.path.exists(filename):
            backup_name = filename + ".corrupted"
            shutil.move(filename, backup_name)
            print(f"Backed up corrupted file to: {backup_name}")
            
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        return wb


def add_paper_to_workbook(wb, paper_data):
    """
    Adds a new sheet to the workbook for the given paper data.
    """
    # Construct base sheet name: Author_Year
    # Extract first author surname
    authors = paper_data.get("Authors", "Unknown")
    if isinstance(authors, list):
         first_author = authors[0].split()[-1] if authors else "Unknown"
    else:
         # simple heuristic if it's a string
         first_author = authors.split(',')[0].split()[-1] if ',' in authors else authors.split()[-1]
    
    year = str(paper_data.get("Year", "Unknown"))
    
    # Clean up author/year to be safe
    # Clean up author/year
    first_author = "".join(x for x in first_author if x.isalnum())
    year = "".join(x for x in year if x.isalnum())
    
    # Include Short Title for better sorting/identification
    title = paper_data.get("Title", "")
    # Remove non-alphanumeric (keep spaces for readability then remove later? or just strict alnum)
    # The clean_sheet_name handles invalid chars, but let's be cleaner for the base
    short_title = "".join(x for x in title if x.isalnum())
    
    # Construct base: Author_Year_Title
    # We need to be careful with length. 31 limit.
    # heuristic: Author (max 10) + _ + Year (4) + _ + Title (remaining)
    # This ensures Author_Year is preserved.
    
    author_part = first_author[:10]
    year_part = year[:4]
    
    # Author_Year_ is roughly 10+1+4+1 = 16 chars. 
    # Remaining for title = 31 - 16 = 15 chars.
    
    base_name = f"{author_part}_{year_part}_{short_title}"
    
    # get_unique_sheet_name will truncate the whole thing to 31 if needed
    sheet_name = get_unique_sheet_name(wb, base_name)
    
    ws = wb.create_sheet(title=sheet_name)
    
    # Headers
    ws['A1'] = "Category/Question"
    ws['B1'] = "Extracted Answer"
    
    # Define the order of keys to write
    keys_order = [
        "Title", "Authors", "Journal", "Volume", "Pages", "Year", "DOI",
        "Central Problem", "Central Hypothesis", "Central Objective",
        "Central Independent Variables", "Central Dependent Variables",
        "Methodology & Tools", "Central Result", "Central Conclusion",
        "Short Summary", "Glossary"
    ]
    
    row_idx = 2
    for key in keys_order:
        val = paper_data.get(key, "N/A")
        
        # Consistent Glossary Formatting
        if key == "Glossary":
            import ast
            formatted_glossary = []
            
            # Helper to extract Term/Def
            def extract_term_def(item):
                if isinstance(item, dict):
                     t = item.get("Term", item.get("term", ""))
                     d = item.get("Definition", item.get("definition", ""))
                     return t, d
                return None, None

            try:
                # Case 1: List of dicts
                if isinstance(val, list):
                    for item in val:
                        t, d = extract_term_def(item)
                        if t:
                            formatted_glossary.append(f"• {t}: {d}")
                        else:
                            # fallback if just string list?
                            if isinstance(item, str):
                                formatted_glossary.append(f"• {item}")

                # Case 2: String representation of list
                elif isinstance(val, str) and val.strip().startswith("["):
                    glossary_list = ast.literal_eval(val)
                    if isinstance(glossary_list, list):
                        for item in glossary_list:
                             t, d = extract_term_def(item)
                             if t:
                                 formatted_glossary.append(f"• {t}: {d}")
                
                # Case 3: Already formatted string or other string
                elif isinstance(val, str):
                    formatted_glossary.append(val)
                    
            except Exception as e:
                # If parsing fails, just dump str
                formatted_glossary.append(str(val))
            
            if formatted_glossary:
                val = "\n".join(formatted_glossary)
            else:
                val = "N/A"
        
        elif isinstance(val, list):
            val = ", ".join(str(v) for v in val)
        else:
            val = str(val)
            
        ws.cell(row=row_idx, column=1, value=key).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=row_idx, column=2, value=val)
        row_idx += 1
        
    # Formatting
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 100
    
    # Global Alignment & Wrapping
    for row in ws.iter_rows(min_row=1, max_row=row_idx-1, min_col=1, max_col=2):
        for cell in row:
            # Top-left alignment for readability, especially for long text
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
    # Center align headers
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=False)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=False)
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=12)
    ws['B1'].font = openpyxl.styles.Font(bold=True, size=12)

def sort_sheets_alphabetically(wb):
    """Sorts all sheets in the workbook alphabetically, EXCEPT special sheets."""
    special_sheets = ["Knowledge Graph", "Dashboard"]
    
    # Get all paper sheets
    paper_sheets = [n for n in wb.sheetnames if n not in special_sheets]
    paper_sheets.sort()
    
    # Define desired order
    final_order = []
    
    # 1. Knowledge Graph (if exists)
    if "Knowledge Graph" in wb.sheetnames:
        final_order.append("Knowledge Graph")
        
    # 2. Dashboard (if exists)
    if "Dashboard" in wb.sheetnames:
        final_order.append("Dashboard")
        
    # 3. Papers
    final_order.extend(paper_sheets)
    
    # Reorder
    # This loop works by moving each sheet to its correct relative position
    for i, name in enumerate(final_order):
        wb.move_sheet(wb[name], offset=i - wb.index(wb[name]))

def cleanup_empty_sheets(wb):
    """Removes 'Sheet' or 'Sheet1' if they are default/empty."""
    for default_name in ["Sheet", "Sheet1"]:
        if default_name in wb.sheetnames:
            ws = wb[default_name]
            if ws.max_row <= 1: # Empty or just header (if generated)?
                # Usually default sheet is empty (max_row=1 if accessed or 0)
                # Let's just delete it if we have other sheets
                if len(wb.sheetnames) > 1:
                    del wb[default_name]

def save_workbook(wb, filename):
    cleanup_empty_sheets(wb)
    update_dashboard(wb) # Regenerate dashboard content
    sort_sheets_alphabetically(wb) # Set final order (Graph, DB, Papers)
    wb.save(filename)

def update_dashboard(wb):
    """Regenerates the Dashboard sheet with TOC."""
    DASHBOARD_NAME = "Dashboard"
    
    # Create or Get Dashboard
    if DASHBOARD_NAME in wb.sheetnames:
        ws = wb[DASHBOARD_NAME]
        # Clear existing data
        ws.delete_rows(1, ws.max_row + 1)
    else:
        # Create
        ws = wb.create_sheet(DASHBOARD_NAME) 
    
    # NOTE: Positioning is handled by sort_sheets_alphabetically now.

    # Headers
    headers = ["Sheet Name", "Title", "Authors", "Year", "Short Summary", "Glossary"]
    ws.append(headers)
    
    # Style Headers
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Iterate sheets
    row_idx = 2
    for sheet_name in wb.sheetnames:
        if sheet_name == DASHBOARD_NAME:
            continue
            
        sheet = wb[sheet_name]
        
        # Extract data from the sheet content. 
        data_map = {}
        # Scan first, say, 25 rows (Glossary might be further down)
        for r in range(1, 30):
            k = sheet.cell(row=r, column=1).value
            v = sheet.cell(row=r, column=2).value
            if k and v:
                data_map[str(k).strip()] = v
        
        title = data_map.get("Title", sheet_name)
        authors = data_map.get("Authors", "Unknown")
        year = data_map.get("Year", "Unknown")
        summary = data_map.get("Short Summary", data_map.get("Central Result", "N/A"))
        
        # Glossary formatting
        glossary_raw = data_map.get("Glossary", "N/A")
        glossary_text = str(glossary_raw)
        
        # Attempt to make it prettier if it looks like a list/string representation of list
        # We don't want to import json here just for this if we can avoid complex parsing logic inside a writer,
        # but basic cleanup helps.
        # If it was saved as string representation of list of dicts: "[{'Term': 'A', 'Definition': 'B'}, ...]"
        import ast
        try:
            if isinstance(glossary_raw, str) and glossary_raw.strip().startswith("["):
                 # It might be a stringified list from the main sheet write
                 glossary_list = ast.literal_eval(glossary_raw)
                 if isinstance(glossary_list, list):
                     formatted_items = []
                     for item in glossary_list:
                         if isinstance(item, dict):
                             term = item.get("Term", item.get("term", ""))
                             defn = item.get("Definition", item.get("definition", ""))
                             if term:
                                 formatted_items.append(f"{term}: {defn}")
                     if formatted_items:
                         glossary_text = "\n".join(formatted_items)
        except:
             pass # Keep raw string if parsing fails
             
        # Write to Dashboard
        # A: Sheet Name (Hyperlink)
        c_sheet = ws.cell(row=row_idx, column=1, value=sheet_name)
        c_sheet.hyperlink = f"#'{sheet_name}'!A1"
        c_sheet.style = "Hyperlink"
        
        # B: Title
        ws.cell(row=row_idx, column=2, value=title)
        
        # C: Authors
        ws.cell(row=row_idx, column=3, value=authors)
        
        # D: Year
        ws.cell(row=row_idx, column=4, value=year)
        
        # E: Short Summary
        ws.cell(row=row_idx, column=5, value=summary)
        
        # F: Glossary
        ws.cell(row=row_idx, column=6, value=glossary_text)
        
        row_idx += 1
        
    # Formatting
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 60
    
    # Alignment
    for row in ws.iter_rows(min_row=2, max_row=row_idx-1, min_col=1, max_col=6):
        for cell in row:
             cell.alignment = Alignment(vertical='center', wrap_text=True)

    # AutoFilter
    ws.auto_filter.ref = ws.dimensions
